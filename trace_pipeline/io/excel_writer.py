"""Excel 写入 — 多工作表输出（每区一个 sheet）。"""
from __future__ import annotations

import logging
import math
import numbers
from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..analysis.models import NodeAnalysis
from ..geology.statistics import TraceStatistics
from ..models import TraceData

logger = logging.getLogger(__name__)

__all__ = [
    "DEFAULT_LAYOUT",
    "ExcelLayout",
    "ExcelSection",
    "build_excel_sections",
    "build_result_workbook_sections",
    "write_excel_multi_sheets",
]

_SUMMARY_TITLES = {"基本信息", "裂隙情况", "计算数据"}


@dataclass(frozen=True)
class ExcelSection:
    """单个 Excel 输出区段。"""

    df: pd.DataFrame
    startrow: int
    startcol: int
    header: bool
    title: str = ""


@dataclass(frozen=True)
class ExcelLayout:
    """生成 Excel 工作表的布局规格。"""

    base_info_row: int = 0
    data_gap: int = 1
    raw_col_start: int = 0
    rot_col_start: int = 6
    orient_col_start: int = 12
    column_width: int = 14
    min_column_width: int = 10
    max_column_width: int = 28
    summary_min_width: int = 12
    summary_max_width: int = 16
    gap_column_width: int = 3
    raw_column_width: int = 12
    rotated_column_width: int = 14
    orientation_column_width: int = 12
    segment_length_column_width: int = 16
    trace_type_column_width: int = 10


DEFAULT_LAYOUT = ExcelLayout()


def _is_summary_section(section: ExcelSection) -> bool:
    return section.title in _SUMMARY_TITLES


_CJK_FONT = "SimSun"
_WESTERN_FONT = "Times New Roman"


def _is_cjk(char: str) -> bool:
    """判断单个字符是否属于 CJK（中文）Unicode 区块。"""
    cp = ord(char)
    return (
        (0x4E00 <= cp <= 0x9FFF)      # CJK 统一汉字
        or (0x3400 <= cp <= 0x4DBF)   # 扩展 A
        or (0x20000 <= cp <= 0x2A6DF) # 扩展 B
        or (0xF900 <= cp <= 0xFAFF)   # 兼容汉字
        or (0x2F800 <= cp <= 0x2FA1F) # 兼容补充
        or (0x3000 <= cp <= 0x303F)   # CJK 符号与标点
        or (0xFF00 <= cp <= 0xFFEF)   # 全角字符
        or (0x2E80 <= cp <= 0x2FFF)   # 偏旁部首 / 康熙部首 / 表意描述符
        or (0x31C0 <= cp <= 0x31EF)   # CJK 笔画
    )


def _classify_text(text: str) -> str:
    """返回 'cjk' / 'latin' / 'mixed'。"""
    has_cjk = False
    has_other = False
    for ch in text:
        if _is_cjk(ch):
            has_cjk = True
        else:
            has_other = True
        if has_cjk and has_other:
            return "mixed"
    return "cjk" if has_cjk else "latin"


def _build_mixed_font_text(
    value: str,
    bold: bool = False,
    color: str | None = None,
) -> CellRichText:
    """将混合中英文文本拆为 CellRichText，中文=宋体，英文/数字=Times New Roman。"""
    blocks: list[TextBlock] = []
    current_text = ""
    current_cjk: bool | None = None
    for ch in value:
        is_cjk = _is_cjk(ch)
        if current_cjk is None:
            current_cjk = is_cjk
            current_text = ch
        elif is_cjk == current_cjk:
            current_text += ch
        else:
            blocks.append(_make_text_block(current_text, current_cjk, bold, color))
            current_text = ch
            current_cjk = is_cjk
    if current_text:
        assert current_cjk is not None
        blocks.append(_make_text_block(current_text, current_cjk, bold, color))
    return CellRichText(*blocks)


def _make_text_block(
    text: str,
    is_cjk: bool,
    bold: bool,
    color_str: str | None,
) -> TextBlock:
    font_kwargs: dict[str, object] = {"rFont": _CJK_FONT if is_cjk else _WESTERN_FONT}
    if bold:
        font_kwargs["b"] = True
    if color_str:
        font_kwargs["color"] = Color(rgb="00" + color_str)
    return TextBlock(InlineFont(**font_kwargs), text)


def _apply_cell_font(cell, *, bold: bool = False, color: str | None = None) -> None:
    """按单元格内容类型设置字体：中文→宋体，数字/英文→Times New Roman。"""
    value = cell.value
    if value is None:
        return
    if isinstance(value, str):
        classification = _classify_text(value)
        if classification == "mixed":
            cell.value = _build_mixed_font_text(value, bold=bold, color=color)
        else:
            font_name = _CJK_FONT if classification == "cjk" else _WESTERN_FONT
            font_kwargs: dict = {"name": font_name, "bold": bold}
            if color:
                font_kwargs["color"] = color
            cell.font = Font(**font_kwargs)
    elif isinstance(value, (numbers.Integral, numbers.Real)):
        cell.font = Font(name=_WESTERN_FONT, bold=bold)


def _round_float(value: float, digits: int = 4) -> float | None:
    value = float(value)
    if not math.isfinite(value):
        return None
    return round(value, digits)


def _format_excel_cell_value(value, unit: str = "") -> str:
    if value is None:
        text = "N/A"
    elif isinstance(value, numbers.Integral):
        text = str(int(value))
    elif isinstance(value, numbers.Real):
        text = f"{float(value):.4f}".rstrip("0").rstrip(".")
    else:
        text = str(value)
    if unit == "°":
        return f"{text}{unit}"
    return f"{text} {unit}".strip()


def _one_row_df(items: Sequence[tuple[str, str]]) -> pd.DataFrame:
    return pd.DataFrame(
        [[value for _label, value in items]],
        columns=[label for label, _value in items],
    )


def _source_tag(source: str) -> str:
    """来源标注短标签，用于 Excel/表格。"""
    mapping = {
        "measured": "M",
        "window": "W",
        "window_equivalent": "W",
        "hull": "E",
        "endpoint": "E",
        "segment": "E",
        "estimated": "est",
    }
    tag = mapping.get(source)
    return f" ({tag})" if tag else ""


def _build_summary_sections(
    trace: TraceData,
    statistics: TraceStatistics | None,
    layout: ExcelLayout,
) -> list[ExcelSection]:
    mean_length = statistics.mean_trace_length if statistics is not None else trace.mean_length
    basic_items = [
        ("测线走向", _format_excel_cell_value(round(trace.scanline_azimuth, 2), "°")),
        (
            "测线长度",
            _format_excel_cell_value(_round_float(statistics.scanline_length), "m")
            if statistics is not None
            else "N/A",
        ),
        ("平均迹长", _format_excel_cell_value(_round_float(mean_length), "m")),
        (
            "露头面积",
            (
                _format_excel_cell_value(_round_float(statistics.outcrop_area), "m²")
                + _source_tag(statistics.outcrop_area_source)
            )
            if statistics is not None
            else "N/A",
        ),
    ]
    sections: list[ExcelSection] = [
        ExcelSection(
            df=_one_row_df(basic_items),
            startrow=layout.base_info_row,
            startcol=layout.raw_col_start,
            header=True,
            title="基本信息",
        ),
    ]

    if statistics is None:
        return sections

    fracture_items = [
        ("迹线数量", _format_excel_cell_value(statistics.total_count)),
        ("I型裂隙数", _format_excel_cell_value(statistics.type_i_count)),
        ("II型裂隙数", _format_excel_cell_value(statistics.type_ii_count)),
        ("III型裂隙数", _format_excel_cell_value(statistics.type_iii_count)),
    ]
    calculation_items = [
        ("线密度(P₁₀)", _format_excel_cell_value(_round_float(statistics.p10), "m⁻¹")),
        (
            "面密度(P₂₀)",
            _format_excel_cell_value(_round_float(statistics.p20), "m⁻²")
            + _source_tag(statistics.p20_source),
        ),
        (
            "面累计长度密度(P₂₁)",
            _format_excel_cell_value(_round_float(statistics.p21), "m⁻¹")
            + _source_tag(statistics.p21_source),
        ),
        ("有效取样窗数量", _format_excel_cell_value(statistics.valid_window_count)),
    ]
    if statistics.window_validation_warning:
        calculation_items.append(
            ("校验告警", statistics.window_validation_warning),
        )
    sections.extend([
        ExcelSection(
            df=_one_row_df(fracture_items),
            startrow=layout.base_info_row,
            startcol=layout.rot_col_start,
            header=True,
            title="裂隙情况",
        ),
        ExcelSection(
            df=_one_row_df(calculation_items),
            startrow=layout.base_info_row,
            startcol=layout.orient_col_start,
            header=True,
            title="计算数据",
        ),
    ])
    return sections


def _section_row_count(section: ExcelSection) -> int:
    return (1 if section.title else 0) + (1 if section.header else 0) + len(section.df)


def build_result_workbook_sections(
    trace: TraceData,
    rotated_xy: np.ndarray,
    statistics: TraceStatistics | None = None,
    node_analysis: NodeAnalysis | None = None,
    layout: ExcelLayout = DEFAULT_LAYOUT,
) -> list[ExcelSection]:
    """构建多工作表导出的 DataFrame 区段（含节点统计）。"""
    if rotated_xy.shape != trace.endpoints.shape:
        raise ValueError(
            f"旋转坐标形状 {rotated_xy.shape} 与原始坐标 {trace.endpoints.shape} 不一致"
        )
    if not np.isfinite(rotated_xy).all():
        raise ValueError("旋转坐标包含 NaN 或 inf")

    summary_sections = _build_summary_sections(trace, statistics, layout)

    raw_df = pd.DataFrame(
        trace.endpoints,
        columns=["起点X", "起点Y", "终点X", "终点Y"],
    )

    rot_df = pd.DataFrame(
        rotated_xy,
        columns=["旋转后起点X", "旋转后起点Y", "旋转后终点X", "旋转后终点Y"],
    )

    orient_df = pd.DataFrame({
        "节理走向(°)": np.round(trace.joint_strikes, 2),
        "端点距离": np.round(trace.lengths, 4),
        "测段长度(r5+r7)": np.round(trace.segment_lengths, 4),
    })
    if statistics is not None:
        if len(statistics.trace_types) != trace.count:
            raise ValueError(
                f"迹线类型数量 {len(statistics.trace_types)} 与迹线数量 {trace.count} 不一致"
            )
        orient_df["迹线类型"] = list(statistics.trace_types)

    sections: list[ExcelSection] = [
        *summary_sections,
        ExcelSection(raw_df, 0, 0, True, "原始端点坐标"),
        ExcelSection(rot_df, 0, 0, True, "旋转后端点坐标"),
        ExcelSection(orient_df, 0, 0, True, "走向与长度"),
    ]

    if node_analysis is not None:
        sections.extend(_build_node_sections(node_analysis, trace))

    return sections


# 保留旧别名
build_excel_sections = build_result_workbook_sections


def _build_node_sections(
    node_analysis: NodeAnalysis,
    trace: TraceData,
) -> list[ExcelSection]:
    """构建节点相关的 Excel 区段。"""
    tc = node_analysis.type_counts
    stats_items = [
        ("节点总数", str(node_analysis.node_count)),
        ("自由端点(I)", str(tc.get("I", 0))),
        ("三叉节点(Y)", str(tc.get("Y", 0))),
        ("相交节点(X)", str(tc.get("X", 0))),
        ("重叠节点", str(tc.get("overlap", 0))),
        ("多交汇点", str(tc.get("multi", 0))),
        ("交点事件数", str(node_analysis.intersection_count)),
        ("节点密度", _format_excel_cell_value(node_analysis.node_density(trace.measured_outcrop_area), "个/m²")),
        ("合并容差", _format_excel_cell_value(node_analysis.nodes[0].x if node_analysis.nodes else None, "")),
        ("跳过退化线段数", str(node_analysis.degenerate_skipped)),
    ]
    stats_df = _one_row_df(stats_items)

    detail_records = []
    for node in node_analysis.nodes:
        detail_records.append({
            "节点ID": node.node_id,
            "X": round(node.x, 4),
            "Y": round(node.y, 4),
            "类型": node.type_label,
            "度数": node.degree,
            "连接迹线": ",".join(str(i) for i in node.trace_indices),
            "事件数": node.event_count,
            "含端点": "是" if node.is_endpoint else "否",
            "含相交": "是" if node.is_intersection else "否",
            "含重叠": "是" if node.is_overlap else "否",
        })
    detail_df = pd.DataFrame(detail_records)

    inter_records = []
    for ev in node_analysis.intersections:
        inter_records.append({
            "迹线A": ev.trace_a,
            "迹线B": ev.trace_b,
            "交点X": round(ev.x, 4),
            "交点Y": round(ev.y, 4),
            "参数t": round(ev.t, 4),
            "参数u": round(ev.u, 4),
            "事件类型": "相交" if ev.kind == "internal" else ("端点接触" if ev.kind == "endpoint" else "重叠"),
        })
    inter_df = pd.DataFrame(inter_records) if inter_records else pd.DataFrame(columns=["迹线A", "迹线B", "交点X", "交点Y", "参数t", "参数u", "事件类型"])

    return [
        ExcelSection(stats_df, 0, 0, True, "节点统计"),
        ExcelSection(detail_df, 0, 0, True, "节点明细"),
        ExcelSection(inter_df, 0, 0, True, "节点交点"),
    ]


# ── 多工作表写入（新格式：每个分区一个 sheet）─────────────────────────

def _write_sheet_title(ws, title: str, column_count: int) -> None:
    """在 sheet 首行写入合并标题。"""
    if not title:
        return
    cell = ws.cell(row=1, column=1, value=title)
    _apply_cell_font(cell, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4F81BD")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if column_count > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)


def _style_sheet(ws, df: pd.DataFrame, has_title: bool = True) -> None:
    """对整个 sheet 应用样式。"""
    title_offset = 1 if has_title else 0
    header_row = title_offset + 1
    first_data_row = header_row + 1
    last_row = header_row + len(df)
    first_col = 1
    last_col = df.shape[1]
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 表头样式
    for row in ws.iter_rows(min_row=header_row, max_row=header_row, min_col=first_col, max_col=last_col):
        for cell in row:
            _apply_cell_font(cell, bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
    ws.row_dimensions[header_row].height = 28

    # 数据行样式
    for row in ws.iter_rows(min_row=first_data_row, max_row=last_row, min_col=first_col, max_col=last_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            _apply_cell_font(cell, bold=False)
            if isinstance(cell.value, numbers.Integral):
                cell.number_format = "0"
            elif isinstance(cell.value, numbers.Real):
                cell.number_format = "0.0000"

    # 列宽
    for col_idx in range(first_col, last_col + 1):
        max_width = 10
        for column in ws.iter_cols(min_col=col_idx, max_col=col_idx, values_only=False):
            for cell in column:
                if cell.value is not None:
                    max_width = max(max_width, len(str(cell.value)))
        width = min(28, max(10, max_width * 1.1 + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 冻结窗格（冻结标题行+表头行）
    ws.freeze_panes = f"A{header_row + 1}"


def write_excel_multi_sheets(
    excel_path: str,
    sections: Sequence[ExcelSection],
) -> None:
    """将每个分区写入独立工作表（多 sheet 格式）。"""
    output_dir = Path(excel_path).parent
    if str(output_dir):
        output_dir.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        for section in sections:
            sheet_name = section.title or "数据"
            # DataFrame 从第2行开始写入（第1行留给标题）
            title_offset = 1 if section.title else 0
            section.df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                header=section.header,
                startrow=title_offset,
                startcol=0,
            )
            ws = writer.sheets[sheet_name]
            _write_sheet_title(ws, section.title, section.df.shape[1])
            _style_sheet(ws, section.df, has_title=bool(section.title))

    logger.debug("Excel 多工作表写入完成: %s", excel_path)


